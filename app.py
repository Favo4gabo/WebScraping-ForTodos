# from flask import Flask, request, render_template_string
# import openpyxl
# import pandas
# import requests
# from bs4 import BeautifulSoup

# # Crear una instancia de la aplicación Flask
# app = Flask(__name__)

# # Ruta para la página principal que muestra el formulario
# @app.route('/')
# def index():
#     # Renderiza el formulario HTML
#     return render_template_string(open('index.html').read())


# # Ruta para procesar los datos enviados desde el formulario (automatico)
# @app.route('/automatico', methods=['POST'])
# def automatico():
#     # Obtener las URLs ingresadas en el formulario y dividirlas en una lista
#     file = request.files['file']
#     # urls = request.form['urls'].split('\n')
    
#     # Guardar un archivo
#     # file.save(file.filename)
    
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook['Links']  # Cambia el nombre de la hoja según sea necesario

#     data = []

#     # Recorre las filas y extrae los hipervínculos
#     for row in sheet.iter_rows():
#         cell = row[2]  # Cambia el índice de columna según sea necesario
#         if cell.hyperlink:            
#             urls = cell.hyperlink.target
#             # nombre = cell.value
                                        
#             # # Agregar las urls extraídas a la lista            
#             data.append(urls)

#     print(data)
    
#     for url in data:
#         response = requests.get(url)  # Realizar la solicitud HTTP a la URL
#         soup = BeautifulSoup(response.content, 'html.parser')  # Parsear el contenido HTML
        
#         # Extraer el nombre y el precio del producto usando selectores CSS
#         # Ajusta los selectores según la estructura de la página de Mercado Libre
#         nombre_tag = soup.select_one('h1')
#         if nombre_tag is not None:
#             nombre = nombre_tag.text.strip()
#         precio_tag = soup.select_one('span.andes-money-amount__fraction')
#         if precio_tag is not None:
#             precio = precio_tag.text.strip()
#         condicion_tag = soup.select_one('span.ui-pdp-subtitle')
#         if condicion_tag is not None:
#             condicion = condicion_tag.text.strip()
        
#         # Agregar los datos extraídos a la lista
#         data.append({'Nombre del producto': nombre, 'Precio': precio, 'Condicion':condicion})
#         print(data)

#     # Crear un DataFrame de pandas con los datos extraídos
#     df = pandas.DataFrame(data)
#     # Guardar el DataFrame en un archivo Excel
#     df.to_excel('productos_2°Intento.xlsx', index=False)

#     # Devolver un mensaje de éxito
#     return 'Archivo Excel generado con éxito (Automatico).'


# # Ruta para procesar los datos enviados desde el formulario
# @app.route('/manual', methods=['POST'])
# def manual():
#     # Obtener las URLs ingresadas en el formulario y dividirlas en una lista
#     urls = request.form['urls'].split('\n')
#     data = []

#     # Iterar sobre cada URL para extraer la información del producto
#     for url in urls:
#         response = requests.get(url.strip())  # Realizar la solicitud HTTP a la URL
#         soup = BeautifulSoup(response.content, 'html.parser')  # Parsear el contenido HTML
        
#         # Extraer el nombre y el precio del producto usando selectores CSS
#         # Ajusta los selectores según la estructura de la página de Mercado Libre
#         nombre = soup.select_one('h1').text if soup.select_one('h1') else 'N/A'
#         precio = soup.select_one('span.andes-money-amount__fraction').text if soup.select_one('span.andes-money-amount__fraction') else 'N/A'
#         condicion = soup.select_one("span.ui-pdp-subtitle").text if soup.select_one("span.ui-pdp-subtitle") else 'N/A'
        
#         # Agregar los datos extraídos a la lista
#         data.append({'Nombre del producto': nombre, 'Precio': precio, "Condicion": condicion})

#     # Crear un DataFrame de pandas con los datos extraídos
#     df = pandas.DataFrame(data)
#     # Guardar el DataFrame en un archivo Excel
#     df.to_excel('productos.xlsx', index=False)

#     # Devolver un mensaje de éxito
#     return 'Archivo Excel generado con éxito (Manual).'
    
# if __name__ == '__main__':
#     app.run(debug=True)
    


# # ------------------------------------------------------------------------------
# # CODIGO POR PARTES
# # ------------------------------------------------------------------------------

# ----------------------------------------------------------------------------
# ESTE FUNCIONA
# -------------------------------------------------------------------------------

# from flask import Flask, request, render_template_string
# import openpyxl
# import pandas
# import requests
# from bs4 import BeautifulSoup

# app = Flask(__name__)

# # Ruta para la página principal que muestra el formulario
# @app.route('/')
# def index():
#     # Renderiza el formulario HTML
#     return render_template_string(open('index.html').read())

# # Ruta para procesar los datos enviados desde el formulario
# @app.route('/procesar', methods=['POST'])
# def procesar():
#     # Obtener las URLs ingresadas en el formulario y dividirlas en una lista
#     urls = request.form['urls'].split('\n')
#     data = []

#     # Iterar sobre cada URL para extraer la información del producto
#     for url in urls:
#         response = requests.get(url.strip())  # Realizar la solicitud HTTP a la URL
#         soup = BeautifulSoup(response.content, 'html.parser')  # Parsear el contenido HTML
        
#         # Extraer el nombre y el precio del producto usando selectores CSS
#         # Ajusta los selectores según la estructura de la página de Mercado Libre
#         nombre = soup.select_one('h1').text if soup.select_one('h1') else 'N/A'
#         precio_entero = soup.select_one('span.andes-money-amount__fraction').text if soup.select_one('span.andes-money-amount__fraction') else 'N/A'
#         precio_decimal = soup.select_one('span.andes-money-amount__cents').text if soup.select_one('span.andes-money-amount__cents') else '00'
        
#         # Combinar la parte entera y decimal del precio
#         precio = f"{precio_entero},{precio_decimal}" if precio_entero != 'N/A' else 'N/A'
        
#         # condicion = soup.select_one("span.ui-pdp-subtitle").text if soup.select_one("span.ui-pdp-subtitle") else 'N/A'
        
#         # # # # Solo agregar el producto si la condición es "Usado"
#         # # if condicion == "Usado":
#         # categoria = soup.select_one('a.andes-breadcrumb__item').text if soup.select_one('a.andes-breadcrumb__item') else 'N/A'
#         descripcion_1 = soup.select_one('div.ui-pdp-description').text if soup.select_one('p.ui-pdp-description__content') else 'N/A'
#         descripcion_2 = soup.select_one('div.ui-pdp-collapsable__container').text if soup.select_one('div.ui-pdp-collapsable__container') else 'N/A'
#         descripcion_3 = soup.select_one('p.ui-pdp-description__content').text if soup.select_one('p.ui-pdp-description__content') else 'N/A'    
#         # Extraer URLs de las imágenes
#         imagenes = soup.select('img.ui-pdp-image.ui-pdp-gallery__figure__image')
#         imagen_1 = imagenes[0].get('data-zoom', imagenes[0].get('src')) if len(imagenes) > 0 else 'N/A'
#         imagen_2 = imagenes[1].get('data-zoom', imagenes[1].get('src')) if len(imagenes) > 1 else 'N/A'

#         # Agregar los datos extraídos a la lista
#         # data.append({'Nombre del producto': nombre, 'Precio': precio, "Condicion": condicion, 'Categoria': categoria, 'Descripcion': descripcion})
#         data.append({'Nombre': nombre, 'Descripcion':descripcion_1, 'Descripcion 2':descripcion_2, 'Imagen 1': imagen_1, 'Imagen 2': imagen_2, 'Precio': precio})

#     # Crear un DataFrame de pandas con los datos extraídos
#     df = pandas.DataFrame(data)
#     # Guardar el DataFrame en un archivo Excel
#     df.to_excel('Imagen.xlsx', index=False)
    
#     print(data)

#     # Devolver un mensaje de éxito
#     return 'Archivo Excel generado con éxito.'

# # Ruta para extraer hipervínculos de un archivo Excel y guardarlos en otro archivo Excel
# @app.route('/extract_links', methods=['POST'])
# def extract_links():
#     # Obtener las URLs ingresadas en el formulario y dividirlas en una lista
#     file = request.files['file']
    
#     # Abre el libro de Excel
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook['Links']  # Cambia el nombre de la hoja según sea necesario

#     data = []

#     # Recorre las filas y extrae los hipervínculos
#     for row in sheet.iter_rows():
#         cell = row[2]  # Cambia el índice de columna según sea necesario
#         if cell.hyperlink:
#             nombre = cell.value
#             urls = cell.hyperlink.target

#             # Agregar los datos extraídos a la lista
#             data.append({'Nombre del producto': nombre, 'URL': urls})

#     # Crear un DataFrame de pandas con los datos extraídos
#     df = pandas.DataFrame(data)
#     # Guardar el DataFrame en un archivo Excel
#     df.to_excel('URLs.xlsx', index=False)
    
#     print(data)

#     return "Nombres y URLs extraídos y guardados en 'URLs.xlsx'"

# # Ejecutar la aplicación Flask
# if __name__ == '__main__':
#     app.run(debug=True)
    

    
# from flask import Flask, request, render_template_string
# import openpyxl
# import pandas
# import requests
# from bs4 import BeautifulSoup
# import time

# app = Flask(__name__)

# # Ruta para la página principal que muestra el formulario
# @app.route('/')
# def index():
#     # Renderiza el formulario HTML
#     return render_template_string(open('index.html').read())

# # Ruta para procesar los datos enviados desde el formulario
# @app.route('/procesar', methods=['POST'])
# def procesar():
#     # Obtener las URLs ingresadas en el formulario y dividirlas en una lista
#     urls = request.form['urls'].split('\n')
#     data = []

#     # Iterar sobre cada URL para extraer la información del producto
#     for url in urls:
#         try:
#             response = requests.get(url.strip())  # Realizar la solicitud HTTP a la URL
#             response.raise_for_status()  # Verificar si la solicitud fue exitosa
#             soup = BeautifulSoup(response.content, 'html.parser')  # Parsear el contenido HTML
            
#             # Extraer el nombre y el precio del producto usando selectores CSS
#             nombre = soup.select_one('h1').text if soup.select_one('h1') else 'N/A'
#             precio_entero = soup.select_one('span.andes-money-amount__fraction').text if soup.select_one('span.andes-money-amount__fraction') else 'N/A'
#             precio_decimal = soup.select_one('span.andes-money-amount__cents').text if soup.select_one('span.andes-money-amount__cents') else '00'
            
#             # Combinar la parte entera y decimal del precio
#             precio = f"{precio_entero},{precio_decimal}" if precio_entero != 'N/A' else 'N/A'
            
#             # Verificar múltiples selectores para las descripciones
#             descripcion_1 = (soup.select_one('div.ui-pdp-description').text if soup.select_one('div.ui-pdp-description') else
#                              soup.select_one('p.ui-pdp-description__content').text if soup.select_one('p.ui-pdp-description__content') else
#                              'N/A')
#             descripcion_2 = (soup.select_one('div.ui-pdp-collapsable__container').text if soup.select_one('div.ui-pdp-collapsable__container') else
#                              soup.select_one('div.ui-pdp-collapsable__content').text if soup.select_one('div.ui-pdp-collapsable__content') else
#                              'N/A')
#             descripcion_3 = (soup.select_one('p.ui-pdp-description__content').text if soup.select_one('p.ui-pdp-description__content') else
#                              soup.select_one('div.ui-pdp-description__content').text if soup.select_one('div.ui-pdp-description__content') else
#                              'N/A')

#             # Extraer URLs de las imágenes
#             imagenes = soup.select('img.ui-pdp-image.ui-pdp-gallery__figure__image')
#             imagen_1 = imagenes[0].get('data-zoom', imagenes[0].get('src')) if len(imagenes) > 0 else 'N/A'
#             imagen_2 = imagenes[1].get('data-zoom', imagenes[1].get('src')) if len(imagenes) > 1 else 'N/A'

#             # Agregar los datos extraídos a la lista
#             data.append({'Nombre': nombre, 'Descripcion 1': descripcion_1, 'Descripcion 2': descripcion_2, 'Descripcion 3': descripcion_3, 'Imagen 1': imagen_1, 'Imagen 2': imagen_2, 'Precio': precio})

#             # Pausar entre solicitudes para evitar ser bloqueado
#             time.sleep(2)

#         except requests.RequestException as e:
#             print(f"Error al procesar la URL {url.strip()}: {e}")
#             data.append({'Nombre': 'N/A', 'Descripcion 1': 'N/A', 'Descripcion 2': 'N/A', 'Descripcion 3': 'N/A', 'Imagen 1': 'N/A', 'Imagen 2': 'N/A', 'Precio': 'N/A'})
#         except Exception as e:
#             print(f"Error al procesar la URL {url.strip()}: {e}")
#             data.append({'Nombre': 'N/A', 'Descripcion 1': 'N/A', 'Descripcion 2': 'N/A', 'Descripcion 3': 'N/A', 'Imagen 1': 'N/A', 'Imagen 2': 'N/A', 'Precio': 'N/A'})

#     # Crear un DataFrame de pandas con los datos extraídos
#     df = pandas.DataFrame(data)
#     # Guardar el DataFrame en un archivo Excel
#     df.to_excel('Productos3.xlsx', index=False)
    
#     print(data)

#     # Devolver un mensaje de éxito
#     return 'Archivo Excel generado con éxito.'


# # Ruta para extraer hipervínculos de un archivo Excel y guardarlos en otro archivo Excel
# @app.route('/extract_links', methods=['POST'])
# def extract_links():
#     # Obtener las URLs ingresadas en el formulario y dividirlas en una lista
#     file = request.files['file']
    
#     # Abre el libro de Excel
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook['Links']  # Cambia el nombre de la hoja según sea necesario

#     data = []

#     # Recorre las filas y extrae los hipervínculos
#     for row in sheet.iter_rows():
#         cell = row[2]  # Cambia el índice de columna según sea necesario
#         if cell.hyperlink:
#             nombre = cell.value
#             urls = cell.hyperlink.target

#             # Agregar los datos extraídos a la lista
#             data.append({'Nombre del producto': nombre, 'URL': urls})

#     # Crear un DataFrame de pandas con los datos extraídos
#     df = pandas.DataFrame(data)
#     # Guardar el DataFrame en un archivo Excel
#     df.to_excel('URLs.xlsx', index=False)
    
#     print(data)

#     return "Nombres y URLs extraídos y guardados en 'URLs.xlsx'"

# # Ejecutar la aplicación Flask
# if __name__ == '__main__':
#     app.run(debug=True)


from flask import Flask, request, render_template_string
import openpyxl
import pandas
import requests
from bs4 import BeautifulSoup
import time

app = Flask(__name__)

# Ruta para la página principal que muestra el formulario
@app.route('/')
def index():
    return render_template_string(open('index.html').read())

def get_description(soup, attempts=3):
    for _ in range(attempts):
        descripcion_1 = soup.select_one('div.ui-pdp-description').text if soup.select_one('div.ui-pdp-description') else None
        descripcion_2 = soup.select_one('div.ui-pdp-collapsable__container').text if soup.select_one('div.ui-pdp-collapsable__container') else None
        descripcion_3 = soup.select_one('p.ui-pdp-description__content').text if soup.select_one('p.ui-pdp-description__content') else None
        
        if descripcion_1 or descripcion_2 or descripcion_3:
            return descripcion_1, descripcion_2, descripcion_3
        
        time.sleep(1)  # Espera 1 segundo antes de intentar nuevamente
    
    return 'N/A', 'N/A', 'N/A'

# Ruta para procesar los datos enviados desde el formulario
@app.route('/procesar', methods=['POST'])
def procesar():
    urls = request.form['urls'].split('\n')
    data = []

    for url in urls:
        try:
            response = requests.get(url.strip())
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            
            nombre = soup.select_one('h1').text if soup.select_one('h1') else 'N/A'
            precio_entero = soup.select_one('span.andes-money-amount__fraction').text if soup.select_one('span.andes-money-amount__fraction') else 'N/A'
            precio_decimal = soup.select_one('span.andes-money-amount__cents').text if soup.select_one('span.andes-money-amount__cents') else '00'
            precio = f"{precio_entero},{precio_decimal}" if precio_entero != 'N/A' else 'N/A'
            
            descripcion_1, descripcion_2, descripcion_3 = get_description(soup)
            
            imagenes = soup.select('img.ui-pdp-image.ui-pdp-gallery__figure__image')
            imagen_1 = imagenes[0].get('data-zoom', imagenes[0].get('src')) if len(imagenes) > 0 else 'N/A'
            imagen_2 = imagenes[1].get('data-zoom', imagenes[1].get('src')) if len(imagenes) > 1 else 'N/A'

            data.append({'Nombre': nombre, 'Descripcion 1': descripcion_1, 'Descripcion 2': descripcion_2, 'Descripcion 3': descripcion_3, 'Imagen 1': imagen_1, 'Imagen 2': imagen_2, 'Precio': precio})

            time.sleep(2)

        except requests.RequestException as e:
            print(f"Error al procesar la URL {url.strip()}: {e}")
            data.append({'Nombre': 'N/A', 'Descripcion 1': 'N/A', 'Descripcion 2': 'N/A', 'Descripcion 3': 'N/A', 'Imagen 1': 'N/A', 'Imagen 2': 'N/A', 'Precio': 'N/A'})
        except Exception as e:
            print(f"Error al procesar la URL {url.strip()}: {e}")
            data.append({'Nombre': 'N/A', 'Descripcion 1': 'N/A', 'Descripcion 2': 'N/A', 'Descripcion 3': 'N/A', 'Imagen 1': 'N/A', 'Imagen 2': 'N/A', 'Precio': 'N/A'})

    df = pandas.DataFrame(data)
    df.to_excel('Productos4.xlsx', index=False)
    
    print(data)

    return 'Archivo Excel generado con éxito.'

@app.route('/extract_links', methods=['POST'])
def extract_links():
    file = request.files['file']
    workbook = openpyxl.load_workbook(file)
    sheet = workbook['Links']

    data = []

    for row in sheet.iter_rows():
        cell = row[2]
        if cell.hyperlink:
            nombre = cell.value
            urls = cell.hyperlink.target
            data.append({'Nombre del producto': nombre, 'URL': urls})

    df = pandas.DataFrame(data)
    df.to_excel('URLs.xlsx', index=False)
    
    print(data)

    return "Nombres y URLs extraídos y guardados en 'URLs.xlsx'"

if __name__ == '__main__':
    app.run(debug=True)
        
    
